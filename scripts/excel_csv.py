import openpyxl as xl
import csv

PATH = "/Users/gladysbaudet/Desktop/PFE/CDI/24mois_CDI/"
name = "_24m_CDI.xlsx"

def read(name):
    wb = xl.load_workbook(filename=name)
    ws = wb.active
    vect = 683*['']

    ind=0
    for i in range(34,789):

        if (ws['B'+str(i)].value and ws['A'+str(i)].value):
            if (ws['D'+str(i)].value):
                # print(ws['D'+str(i)].value)
                # print(i)
                vect[ind]='Understands'
            ind+=1
    # print(vect)
    return vect
# ______________________________________________________________________________

def fill_headers(name):
    wb = xl.load_workbook(filename=name)
    ws = wb.active
    vect = 683*['']
    ind=0
    for i in range(34,789):

        if (ws['B'+str(i)].value):
            vect[ind] = ws['B'+str(i)].value
            ind += 1
    return vect
# ______________________________________________________________________________

def enum():
    csv_file = open(PATH+'summary.csv', "w")
    writer = csv.writer(csv_file, delimiter=' ', quotechar='|', quoting=csv.QUOTE_MINIMAL)
    # print(fill_headers(PATH+'001'+name))
    # writer.writerow(['lol']*4)
    writer.writerow(fill_headers(PATH+'001'+name))

    # for i in range (1,59):
    for i in range(1,59):
        print(i)
        if i<10:
            nb = '00'+str(i)
        else :
            nb = '0'+str(i)
        vect = read(PATH+nb+name)
    # vect = read(PATH+"001"+name)

        writer.writerow(vect)

enum()
